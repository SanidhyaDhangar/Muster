from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from ..database import get_db
from ..deps import get_current_user, require_roles
from ..models import AttendanceLog, Period, Student
from ..schemas import AttendanceRow, ManualAttendanceRequest
from ..services import already_logged, build_attendance_rows

router = APIRouter(prefix="/api/attendance", tags=["attendance"])

EXPORT_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@router.get("/logs", response_model=list[AttendanceRow])
def attendance_logs(
    branch: str | None = None,
    semester: int | None = None,
    period_id: int | None = None,
    date: str | None = None,
    status: str = "all",
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    return build_attendance_rows(db, branch, semester, period_id, date, status)


@router.post("/manual")
def manual_attendance(payload: ManualAttendanceRequest, db: Session = Depends(get_db), _=Depends(require_roles("admin", "professor"))):
    period = db.get(Period, payload.period_id)
    if period is None:
        raise HTTPException(status_code=404, detail="Period not found")

    day = payload.date or datetime.now().strftime("%Y-%m-%d")
    created = 0
    for roll_no in payload.present_roll_nos:
        student = db.get(Student, roll_no)
        if student is None or already_logged(db, roll_no, period.period_name, day):
            continue
        timestamp = datetime.strptime(f"{day} {datetime.now().strftime('%H:%M:%S')}", "%Y-%m-%d %H:%M:%S")
        db.add(AttendanceLog(
            student_roll_no=student.roll_no,
            student_name=student.name,
            period_name=period.period_name,
            prof_name=period.prof_name,
            timestamp=timestamp,
        ))
        created += 1
    db.commit()
    return {"created": created}


@router.delete("/logs/{log_id}", status_code=204)
def delete_log(log_id: int, db: Session = Depends(get_db), _=Depends(require_roles("admin"))):
    log = db.get(AttendanceLog, log_id)
    if log is None:
        raise HTTPException(status_code=404, detail="Log entry not found")
    db.delete(log)
    db.commit()


@router.get("/export")
def export_attendance(
    branch: str | None = None,
    semester: int | None = None,
    period_id: int | None = None,
    date: str | None = None,
    status: str = "all",
    db: Session = Depends(get_db),
    _=Depends(require_roles("admin", "professor")),
):
    rows = build_attendance_rows(db, branch, semester, period_id, date, status)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Attendance"

    headers = ["Roll No", "Name", "Branch", "Semester", "Period", "Professor", "Status", "Timestamp"]
    sheet.append(headers)
    header_fill = PatternFill(start_color="E8EEF7", end_color="E8EEF7", fill_type="solid")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill

    for row in rows:
        sheet.append([
            row.student_roll_no,
            row.student_name,
            row.branch,
            row.semester,
            row.period_name,
            row.prof_name,
            row.status,
            row.timestamp or "",
        ])

    for column in sheet.columns:
        width = max((len(str(cell.value)) for cell in column if cell.value is not None), default=10)
        sheet.column_dimensions[column[0].column_letter].width = width + 2

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"attendance_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type=EXPORT_MIME,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
